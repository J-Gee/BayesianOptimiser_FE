'''
Parser class contains methods to parse to parse KUKA output and
generate submission files with the points.

directory_path is the class variable pointing to the running directory 
containing folders runqueue, running, completed.

Adapted for the Formulation Engine by Jack Gee
Provides Excel sheets in the HOLLY direct upload format as submission files

'''
import pandas as pd
import os
import re
import datetime
from openpyxl import load_workbook

class Parser:
    FLOAT_FORMAT = "{:.4f}"
    
    def __init__(self, compounds, directory_path):
        '''
        The initializer just takes the list of compound names to work with.
        '''
        self.directory_path = directory_path
        self.processed_files = {} # Dictonary of pandas frames. One for every processed file. If file from another batch_name, there will be None
        self.patterns = {} # Patterns for quantities and sample number subsitutions

        for chem in compounds:
            self.patterns[chem] = re.compile(r"\$\{"+chem+r"\}")
        
        self.patterns['idx'] = re.compile(r"\$\{idx\}")
        self.patterns["batch_name"] = re.compile(r"\$\{batch_name\}")
        self.patterns['sample_number'] = re.compile(r"\$\{sample_number\}")
        
        #Water hack pattern
        self.water = re.compile(r"\$\{water\}")

        #Test hack pattern.
        self.test = re.compile(r"\$\{test\}")

        self.submission_header = "" # fixed string from batch.template to add before experimental compositions
        self.submission_compounds = "" # string from batch.template giving comma-separated list of compounds
        self.submission_line = "" # string from batch.template that would be used for every vial content

        try:
            with open(self.directory_path+'batch.template',"r") as f:

                for line in f.readlines():
                    if line.startswith('$'):
                        self.submission_line = line
                    elif line.startswith('SampleIndex,'):
                        self.submission_compounds = line
                    else:
                        self.submission_header += line

                self.submission_header = self.submission_header.rstrip()
                    
        except IOError:
            print("Template file for submission is missing.\nChange the directory_path variable in the Parser class.")
            print("Present path set to: "+self.directory_path+'batch.template') 
     
    def __repr__(self):
        return str(self.__class__) + ": " + str(self.__dict__)
    
    def process_completed_file(self, filename, experiment_name=None):
        '''
        Generate pandas dataframe for a given file and put it into
        self.processed_files['filename']. It relies on SampleIndex line.

        Returns True if the file belonged to the same experiment

        Note:  (1) I use 'exp_name in first_line' to check whether
                    the file belongs to our experiment. Thus, it is better
                    to use some fixed prefix for ml-driven runs
        '''
        try:
            measurements_line = 0
            path = self.directory_path+'completed/'+filename

            with open(path, "r") as f:
                if experiment_name is not None:
                    line = f.readline()
                    if experiment_name not in line:
                        self.processed_files[filename] = None
                        return False

                for line in f.readlines():
                    if line.startswith('SampleIndex'):
                        break
                    measurements_line += 1
 
            with open(path,"r") as f:
                frame = pd.read_csv(f, skiprows=range(measurements_line))
                self.processed_files[filename] = frame
                return True

        except IOError:
            print("Processed file cannot be found.\nChange the directory_path varriable in the Parser class.")
    
    def process_completed_folder(self, experiment_name=None):
        '''
        Checks whether there are new experimental measurements.
        Returns the list of files that were processed if any.
        '''
        try:
            newly_processed = []
            for f in sorted(os.listdir(self.directory_path+'completed/')):
                if os.path.isfile(self.directory_path+'completed/'+f) and os.path.splitext(f)[1]=='.run':
                    if f not in self.processed_files:
                        if self.process_completed_file(f, experiment_name):
                            newly_processed.append(f)

        except IOError:
            print("Cannot find the completed folder.\nChange the directory_path varriable in the Parser class.") 
        
        return newly_processed
    
    def process_running(self,experiment_name=None):
        '''
        Returns a list of pandas dataframes. 
        A duplicate of process_completed_folder and process_completed_file, for the two running folders.
        '''
        dfs=[]
        # Assemble list of files, and then dataframes for all experiments in running folder
        try:
            running_files = []
            for f in os.listdir(os.path.join(self.directory_path,'running')):
                if os.path.isfile(os.path.join(self.directory_path,'running',f)) and os.path.splitext(f)[1]=='.run': 
                    running_files.append(os.path.join('running',f)) 
        except IOError:
            print("Cannot find the running folder.\nChange the directory_path variable in the Parser class.") 
                    
        # Now for the runqueue    
        try:
            for f in os.listdir(os.path.join(self.directory_path,'runqueue')):
                if os.path.isfile(os.path.join(self.directory_path,'runqueue',f)) and os.path.splitext(f)[1]=='.run': 
                    running_files.append(os.path.join('runqueue',f)) 
        except IOError:
            print("Cannot find the runqueue folder.\nChange the directory_path variable in the Parser class.") 
        
        for file in running_files:
            measurements_line = 0
            path = os.path.join(self.directory_path,file)
            with open(path, "r") as f:
                if experiment_name is not None:
                    line = f.readline()
                    if experiment_name not in line:
                        print("WARNING: runqueue contains files with different experiment name!")
                        continue
                for line in f.readlines():
                    if line.startswith('SampleIndex'): break
                    measurements_line += 1 # finds where headers start and breaks when its found
            with open(path,"r") as f:
                frame = pd.read_csv(f, skiprows=range(measurements_line))
                dfs.append(frame)

        return dfs
    
    def submit_mini_batch(self, batch_name, quantities_list, liquids):
        '''
        Generates a new file for the workflow and puts it in the runqueue folder.
        E.g. quantities_list = [{'P10-HS' : 5, 'AscorbicAcid0-1M' : 4.99}, {'P10-HS' : 5.1, 'AscorbicAcid0-1M' : 4.39}]
        liquids = {'AscorbicAcid0-1M', 'NaCL-1-0M'}
        batch_name should be consistent with FE's direct upload experiment files for HOLLY
        The filename is set to batch_name + '.xlsx'


        '''
        try:

            path = self.directory_path+'/runqueue/' + batch_name + '.xlsx'
            book = load_workbook("fe_batch.xlsx")
            worksheet_names = book.sheetnames

            sub_info = self.submission_header.split("\n")
            sub_info_split = []
            for item in sub_info:
                sub_info_split.append(item.split(",")[0])
                sub_info_split.append(item.split(",")[1])

            sub_info_dict = dict(zip(sub_info_split[::2], sub_info_split[1::2])) #makes dictionary of params in batch.temp

            book.active = worksheet_names.index("Experiment Details") #sets params in exp. details sheet
            ws = book.active
            ws.cell(row=4, column=2).value = batch_name #NEEDS TO BE SAME AS FILE NAME FOR TRACEABILITY
            ws.cell(row=5, column=2).value = sub_info_dict["objective"]
            ws.cell(row=11, column=2).value = sub_info_dict["o2_level"]
            ws.cell(row=12, column=2).value = sub_info_dict["cap_vials"]

            book.active = worksheet_names.index("Formulations") #sets working sheet to formulations to print out amounts
            ws = book.active




            for i, quantity in enumerate(quantities_list):
                # print(i, quantity)
                line = self.submission_line

                #############################################
                # JUST ADD WATER. REMOVE FOR OTHER EXPERIMENTS
                water = 5

                # TEST
                # test = 10

                for chem, amount in quantity.items():
                    if chem in liquids:
                        water = water - amount

                    # TEST
                    # test = test - (0.5 - amount)**2

                if water < 0:
                    print("LAST CHINESE WARNING! The constraints did not work. Total volume is more than 5ml.")
                    water = 0
                #############################################

                for chem, amount in quantity.items():
                    line = self.patterns[chem].sub(self.FLOAT_FORMAT.format(amount), line)

                line = self.patterns['idx'].sub(str(i), line)
                line = self.patterns["batch_name"].sub(batch_name, line)
                line = self.patterns['sample_number'].sub(str(i + 1), line)


                #############################################
                # JUST ADD WATER. REMOVE FOR OTHER EXPERIMENTS
                line = self.water.sub(self.FLOAT_FORMAT.format(water), line)

                # TEST
                # line = self.test.sub(self.FLOAT_FORMAT.format(test), line)
                #############################################

                #f.write(line)
                line = line.split(",")
                headers = self.submission_compounds.split(",")


                line = line[2:]
                headers = headers[2:] #trims id and number as not required.
                row = i+3



                ws.cell(row=i+3, column=1).value = line[0] #puts batch_name+num into form code
                ws.cell(row=i + 3, column=2).value = line[0] #puts batch name+num into form des until HOLLY supports form codes
                #ws.cell(row=i + 3, column=2).value = sub_info_dict["description"]

                ws.cell(row=i + 3, column=3).value = sub_info_dict["hazard_1"]
                ws.cell(row=i + 3, column=4).value = sub_info_dict["hazard_2"]
                ws.cell(row=i + 3, column=5).value = sub_info_dict["hazard_3"]
                col=8 #sets col to H
                for n in range(len(line)):
                    if headers[n] == "Name":
                        continue
                    ws.cell(row=i + 3, column=col).value = headers[n]
                    col += 1
                    ws.cell(row=i + 3, column=col).value = float(line[n])
                    col += 1

                    '''
                    assumes water is always the last addition - which it should be!
                    '''

                    if headers[n] in liquids or headers[n] == headers[-1]:
                        ws.cell(row=i + 3, column=col).value = sub_info_dict["liquid_dispenser"]
                    else:
                        ws.cell(row=i + 3, column=col).value = sub_info_dict["solid_dispenser"]
                    col += 1
                #set end params from col
                ws.cell(row=i + 3, column=col).value = sub_info_dict["orbital_speed_rpm"]
                col += 1
                ws.cell(row=i + 3, column=col).value = sub_info_dict["orbital_id"]
                col += 1
                ws.cell(row=i + 3, column=col).value = sub_info_dict["illumination_time_secs"]
                col += 1
                ws.cell(row=i + 3, column=col).value = sub_info_dict["orbital_id"]
                col += 1
                ws.cell(row=i + 3, column=col).value = sub_info_dict["measurement_method"]

                #f.write('\n')
            book.save(path)
        # try:
        #     path = self.directory_path+'/runqueue/' + batch_name + '.run'
        #     with open(path,"w") as f:
        #         f.write(f'batch_name:{batch_name}\n')
        #
        #
        #
        #
        #         f.write(self.submission_header + '\n\n')
        #
        #         f.write(datetime.datetime.now().strftime('submit_start_datetime:%Y.%m.%d.%H.%M.%S\n\n'))
        #
        #         f.write(self.submission_compounds) #table headers
        #         print(self.submission_compounds)
        #
        #         for i, quantity in enumerate(quantities_list):
        #             # print(i, quantity)
        #             line = self.submission_line
        #              #terms yet to be assigned values
        #             #############################################
        #             #JUST ADD WATER. REMOVE FOR OTHER EXPERIMENTS
        #             water = 5
        #
        #             #TEST
        #             # test = 10
        #
        #             for chem, amount in quantity.items():
        #                 if chem in liquids :
        #                     water = water - amount
        #
        #                 #TEST
        #                 # test = test - (0.5 - amount)**2
        #
        #             if water < 0:
        #                 print("LAST CHINESE WARNING! The constraints did not work. Total volume is more than 5ml.")
        #                 water = 0
        #             #############################################
        #
        #
        #             for chem, amount in quantity.items():
        #                 line = self.patterns[chem].sub(self.FLOAT_FORMAT.format(amount), line) #fills in chemical amounts
        #             line = self.patterns['idx'].sub(str(i), line)
        #              #idx assigned
        #             line = self.patterns["batch_name"].sub(batch_name, line)
        #
        #             line = self.patterns['sample_number'].sub(str(i+1), line)
        #
        #
        #             #############################################
        #             #JUST ADD WATER. REMOVE FOR OTHER EXPERIMENTS
        #             line = self.water.sub(self.FLOAT_FORMAT.format(water), line)
        #
        #             #TEST
        #             # line = self.test.sub(self.FLOAT_FORMAT.format(test), line)
        #             #############################################
        #
        #             f.write(line)
        #
        #             f.write('\n')
 
        except IOError:
            print("Cannot create a batch file.")

if __name__ == "__main__":
    from experiment import Experiment
    exp = Experiment()
    print(exp.rng.keys())
    parser = Parser(list(exp.rng.keys()),'./example_exploratory/')
    parser.process_completed_file("gC3N4-16.run")
    test_submit = [{'P10-HS' : 5, 'AscorbicAcid0-1M' : 4.99}, {'P10-HS' : 5.1, 'AscorbicAcid0-1M' : 4.39}]
    parser.submit_mini_batch(batch_name='test', quantities_list = test_submit)