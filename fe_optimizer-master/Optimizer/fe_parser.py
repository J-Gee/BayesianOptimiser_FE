'''
Parser class contains methods to parse to parse KUKA output and
generate submission files with the points.

directory_path is the class variable pointing to the running directory 
containing folders runqueue, running, completed.

Adapted for the Formulation Engine by Jack Gee
Provides HOLLY exp files





'''
import collections
import csv
import datetime
import os
import re

import pandas
import pandas as pd
from openpyxl import load_workbook
from tkinter import messagebox



########
'''Parser Params'''
SUBSAMPLESIZE= 8
liquid_limit = 1000 # mL
subsample_wash_limit = 35000 # mL
wash_amount = 0

liquid_deadspace = 0.05 # should leave 50 - 100 mL
subsample_deadspace = 0.1 # should leave 4 mL

#wash zeros until labman allow this to be skipped
WASH_ZEROS = False
REMOVE_ZEROS = False
SUBSAMPLE_SWITCH_THROUGH = False
in_wash = []

#delay_time = 10800

##############################################################
class Parser:
    FLOAT_FORMAT = "{:.4f}"

    def __init__(self, compounds, directory_path):
        '''
        The initializer just takes the list of compound names to work with.
        '''
        self.directory_path = directory_path
        self.processed_files = {} # Dictonary of pandas frames. One for every processed file. If file from another batch_name, there will be None
        self.patterns = {} # Patterns for quantities and sample number subsitutions

        for chem in compounds:
            self.patterns[chem] = re.compile(r"\$\{"+chem+r"\}")

        self.patterns['idx'] = re.compile(r"\$\{idx\}")
        self.patterns["batch_name"] = re.compile(r"\$\{batch_name\}")
        self.patterns['sample_number'] = re.compile(r"\$\{sample_number\}")

        #Water hack pattern
        self.water = re.compile(r"\$\{water\}")

        #Test hack pattern.
        self.test = re.compile(r"\$\{test\}")

        self.submission_header = "" # fixed string from batch.template to add before experimental compositions
        self.submission_compounds = "" # string from batch.template giving comma-separated list of compounds
        self.submission_line = "" # string from batch.template that would be used for every vial content

        try:
            with open(self.directory_path+'batch.template',"r") as f:

                for line in f.readlines():
                    if line.startswith('$'):
                        self.submission_line = line
                    elif line.startswith('SampleIndex,'):
                        self.submission_compounds = line
                    else:
                        self.submission_header += line

                self.submission_header = self.submission_header.rstrip()

        except IOError:
            print("Template file for submission is missing.\nChange the directory_path variable in the Parser class.")
            print("Present path set to: "+self.directory_path+'batch.template')

    def __repr__(self):
        return str(self.__class__) + ": " + str(self.__dict__)

    def process_completed_file(self, filename, experiment_name=None):
        '''
        Generate pandas dataframe for a given file and put it into
        self.processed_files['filename']. It relies on SampleIndex line.

        Returns True if the file belonged to the same experiment

        Note:  (1) I use 'exp_name in first_line' to check whether
                    the file belongs to our experiment. Thus, it is better
                    to use some fixed prefix for ml-driven runs


        Jack Note: Opens excel file in comp folder, goes to exp details for title to check its for this exp - could probably just remove this as it requires a 2nd sheet

        '''
        try:
            path = self.directory_path+'completed/'+filename

            book = load_workbook(
                path)  # Basic intialisastion of worksheets into python so we can read off the exp name / #
            worksheet_names = book.sheetnames
            book.active = worksheet_names.index("Output")
            ws = book.active



            #changes with robot due to format of input sheets
            # if experiment_name is not None:
            #     if experiment_name not in ws.cell(row=1, column=1).value:
            #         self.processed_files[filename] = None
            #         return False

            frame = pd.read_excel(path, sheet_name="Output")
            #frame = pd.read_csv(path)
            self.processed_files[filename] = frame

            return True





            #
            # with open(path, "r",filename) as f:
            #     if experiment_name is not None:
            #         line = f.readline()
            #         if experiment_name not in line:
            #             self.processed_files[filename] = None
            #             return False
            #
            #     for line in f.readlines():
            #         if line.startswith('SampleIndex'):
            #             break
            #         measurements_line += 1
            #
            # with open(path,"r") as f:
            #     frame = pd.read_csv(f, skiprows=range(measurements_line))
            #     self.processed_files[filename] = frame
            #     return True

        except IOError:
            print("Processed file cannot be found.\nChange the directory_path varriable in the Parser class.")

    def process_completed_folder(self, experiment_name=None):
        '''
        Checks whether there are new experimental measurements.
        Returns the list of files that were processed if any.
        '''
        try:
            newly_processed = []
            for f in sorted(os.listdir(self.directory_path+'completed/')):
                if os.path.isfile(self.directory_path+'completed/'+f) and os.path.splitext(f)[1]=='.xlsx':
                    if f not in self.processed_files:
                        if self.process_completed_file(f, experiment_name):
                            newly_processed.append(f)

        except IOError:
            print("Cannot find the completed folder.\nChange the directory_path varriable in the Parser class.")

        return newly_processed

    def process_running(self,experiment_name=None):
        '''
        Returns a list of pandas dataframes.
        A duplicate of process_completed_folder and process_completed_file, for the two running folders.
        '''
        dfs=[]
        # Assemble list of files, and then dataframes for all experiments in running folder
        try:
            running_files = []
            for f in os.listdir(os.path.join(self.directory_path,'running')):
                if os.path.isfile(os.path.join(self.directory_path,'running',f)) and os.path.splitext(f)[1]=='.xlsx':
                    running_files.append(os.path.join('running',f))
        except IOError:
            print("Cannot find the running folder.\nChange the directory_path variable in the Parser class.")

        # Now for the runqueue
        try:
            for f in os.listdir(os.path.join(self.directory_path,'runqueue')):
                if os.path.isfile(os.path.join(self.directory_path,'runqueue',f)) and os.path.splitext(f)[1]=='.xlsx':
                    running_files.append(os.path.join('runqueue',f))
        except IOError:
            print("Cannot find the runqueue folder.\nChange the directory_path variable in the Parser class.")

        # I think this segment allows currently running but not complete files to be included in next model?
        for file in running_files:
            measurements_line = 0
            path = os.path.join(self.directory_path,file)
            #print(path)
            book = load_workbook(
                path)  # Basic intialisastion of worksheets into python so we can read off the exp name / #
            worksheet_names = book.sheetnames
            book.active = worksheet_names.index("Experiment Details")
            ws = book.active

            # changes with robot due to format of input sheets
            if experiment_name is not None:
                if experiment_name not in ws.cell(row=4, column=2).value:
                    print("WARNING: runqueue contains files with different experiment name!")
                    continue

            # '''
            # Should point to start of formulations
            # Need the row to end at the cell containing orbital'''
            #
            #
            book.active = worksheet_names.index("Formulations")
            ws = book.active

            data_pulled = False
            headers_pulled = False
            rows_counted = False
            dict_made = False
            overall_dict = {}
            temp_dict = {}
            header_list = ["form_id"]
            header_pos_list = [1]
            values_row = 2
            total_rows = 1
            value_list = []

            while data_pulled == False:
                if headers_pulled == False:
                    col = 8
                    while headers_pulled == False:

                        if ws.cell(row=2, column=col).value == "Flow ID":
                            header_list.append(ws.cell(row=3, column=col).value)
                            header_pos_list.append(col)
                            col += 3
                        else:
                            break

                if rows_counted == False:
                    while rows_counted == False:
                        if ws.cell(row=total_rows, column=1).value is not None:
                            total_rows += 1
                        else:
                            total_rows = total_rows
                            break

                for i in header_pos_list:
                    value_list = []
                    for row in range(total_rows):
                        if row <= values_row: # should tick row to values
                            continue
                        value_list.append(ws.cell(row=row, column=i+1).value)
                    temp_dict = {header_list[header_pos_list.index(i)] : value_list}
                    #overall_dict = overall_dict | temp_dict
                    overall_dict.update(temp_dict)
                break

            frame = pd.DataFrame.from_dict(overall_dict)
            dfs.append(frame)

            #     for head in header_list:
            #
            #         q_point = header_list.index(head)
            #
            #         for i in range(total_rows):
            #             curr_dict.update(head=(ws.cell(row=i+1, column=q_point+1).value))
            #             print(curr_dict)
            #             data_dict = data_dict | curr_dict
            #     data_pulled = True
            #
            # print(data_dict)












            # current_row= []
            # headers_found = False
            # headers_list = []
            # for row in ws.iter_rows(min_row=2):
            #     for i in row:
            #         current_row.append(i.value)
            #     if current_row[0] == None:
            #         break
            #     if headers_found == False:
            #         x = 7
            #         headers_list.append(current_row[x])
            #         while headers_found == False:
            #             x =+ 3
            #             if current_row[x] ==
            #             headers_list.append(current_row[x])






            #frame = pd.read_excel(path, sheet_name="Formulations", skiprows=2)
            #


            # with open(path, "r") as f:
            #     if experiment_name is not None:
            #         line = f.readline()
            #         if experiment_name not in line:
            #             print("WARNING: runqueue contains files with different experiment name!")
            #             continue
            #     for line in f.readlines():
            #         if line.startswith('SampleIndex'): break
            #         measurements_line += 1 # finds where headers start and breaks when its found
            # with open(path,"r") as f:
            #     frame = pd.read_csv(f, skiprows=range(measurements_line))
            #     dfs.append(frame)

        return dfs


    def check_run_queue(self):
        list = []
        list = os.listdir(os.path.join(self.directory_path,'running'))

        return len(list)


    def submit_pandas_batch(self, batch_name, quantities_list, liquids, SUBSAMPLE_SIZE, cats, sr):

        '''
                Generates exp files for use with formulation engine.
                Different from mini batch as it converts line-to-line into pandas dataframes for use with subsampling
                processes

        '''
        try:

            path = self.directory_path + '/runqueue/' + batch_name + '.xlsx'
            book = load_workbook("fe_batch.xlsx")
            worksheet_names = book.sheetnames

            sub_info = self.submission_header.split("\n")
            sub_info_split = []
            for item in sub_info:
                sub_info_split.append(item.split(",")[0])
                sub_info_split.append(item.split(",")[1])

            sub_info_dict = dict(
                zip(sub_info_split[::2], sub_info_split[1::2]))  # makes dictionary of params in batch.temp


            book.active = worksheet_names.index("Experiment Details")  # sets params in exp. details sheet
            ws = book.active
            ws.cell(row=4, column=2).value = batch_name  # NEEDS TO BE SAME AS FILE NAME FOR TRACEABILITY
            ws.cell(row=5, column=2).value = sub_info_dict["objective"]
            ws.cell(row=11, column=2).value = sub_info_dict["o2_level"]
            ws.cell(row=12, column=2).value = sub_info_dict["cap_vials"]



            book.active = worksheet_names.index(
                "Formulations")  # sets working sheet to formulations to print out amounts
            ws = book.active

            #int_working_df = pandas.DataFrame()
            ext_working_df = pandas.DataFrame()
            material_list = pandas.read_csv(self.directory_path + "Material tracking/material_list.csv", delimiter=",")
            material_list = material_list.set_index("Material")
            waste_file = open(self.directory_path + "Material tracking/waste.txt", "r")
            in_wash = waste_file.read()
            in_wash = in_wash.splitlines()
            #print(in_wash)
            waste_file.close()

            #print(quantities_list)

            for i, quantity in enumerate(quantities_list):

                #print(i)
                # print(i, quantity)
                line = self.submission_line
                headers_dict = {}
                headers_dict = collections.defaultdict(list)

                #print(quantities_list)

                #############################################
                # JUST ADD WATER. REMOVE FOR OTHER EXPERIMENTS
                water = 5
                cat_mass = 0
                subsample_amount = 5
                # TEST
                # test = 10

                cat_constr = 3
                sr_constr = 3

                for chem, amount in quantity.items():
                    if chem in liquids:
                        water = water - amount


                    if chem in cats:
                        cat_constr -= 1
                        cat_mass += amount
                        #print(amount)
                    if chem in sr:
                        sr_constr -= 1
                        subsample_amount -= amount


                    # TEST
                    # test = test - (0.5 - amount)**2

                if water < 0:
                    print("LAST CHINESE WARNING! The constraints did not work. Total volume is more than 5ml.")
                    water = 0

                if cat_mass > 5:
                    print("More than 5mg catalyst, " + str(cat_mass) + "mg over")
                if subsample_amount < 0:
                    print("More than 5mL from subsamples")
                # if cat_constr < 0:
                #     print("more than 3 catalysts")
                #
                # if sr_constr < 0:
                #     print("more than 3 SRs")
                #############################################

                for chem, amount in quantity.items():
                    line = self.patterns[chem].sub(self.FLOAT_FORMAT.format(amount), line)

                line = self.patterns['idx'].sub(str(i), line)
                line = self.patterns["batch_name"].sub(batch_name + "-", line)
                line = self.patterns['sample_number'].sub(str(i + 1), line)

                #############################################
                # JUST ADD WATER. REMOVE FOR OTHER EXPERIMENTS
                line = self.water.sub(self.FLOAT_FORMAT.format(water), line)

                # TEST
                # line = self.test.sub(self.FLOAT_FORMAT.format(test), line)
                #############################################

                # f.write(line)
                line = line.split(",")
                headers = self.submission_compounds.split(",")
                line = line[2:]
                headers = headers[2:]  # trims id and number as not required.

                #print(headers)

                #print(working_df)
                row = i + 3

                # ws.cell(row=i + 3, column=1).value = line[0]  # puts batch_name+num into form code
                # ws.cell(row=i + 3, column=2).value = line[
                #     0]  # puts batch name+num into form des until HOLLY supports form codes
                # # ws.cell(row=i + 3, column=2).value = sub_info_dict["description"]
                #
                # ws.cell(row=i + 3, column=3).value = sub_info_dict["hazard_1"]
                # ws.cell(row=i + 3, column=4).value = sub_info_dict["hazard_2"]
                # ws.cell(row=i + 3, column=5).value = sub_info_dict["hazard_3"]
                # col = 8  # sets col to H

                """
                Save the subsample amount for running through multiple batches
                if amount of cat used =! 0 then add delay step? Easy check for new batches?
                """
                wash_counter = 0
                add_mat_b = True

                if i == 13 or i == 14:
                    add_mat_b = False
                #header_new = ''
                for n in range(len(line)):
                    if headers[n] == "Name":
                        continue
                    #ws.cell(row=i + 3, column=col).value = headers[n]
                    # print(headers[n])

                    #curr_material = material_dict[headers[n]]

                    header_new = ''
                    if headers[n] == "water\n":
                        headers[n] = "water"



                    '''
                    Converts opt name for material id for use with holly
                    
                    below adds material amounts to materials_list csv
                    '''

                        # easy pass over these would-be standard samples to not add mat amounts



                    #print(headers[n], material_list.at[headers[n], "type"], str(float(line[n])))
                    #print(material_list)

                    if add_mat_b == True:
                        # print(headers[n])
                        # print(material_list.at[headers[n], "type"])
                        # print("###")

                        if (material_list.at[headers[n], "type"]) == "liquid":
                            #cell_holder = material_list.at[headers[n], "amount1"]
                            if pd.isnull(material_list.at[headers[n], "amount1"]):
                                if i == 29:
                                    material_list.at[headers[n], "amount1"] = (float(line[n])*3)
                                else:
                                    material_list.at[headers[n], "amount1"] = float(line[n])
                                    #print(float(line[n]))
                                #material_list.at[headers[n], "amount1"] = float(line[n])
                                header_new = material_list.at[headers[n], "id1"]
                                #print(header_new)
                            else:
                                if i == 29:
                                    material_list.at[headers[n], "amount1"] = material_list.at[headers[n], "amount1"] + (float(line[n])*3)
                                else: material_list.at[headers[n], "amount1"] = material_list.at[headers[n], "amount1"] + float(line[n])
                                # material_list.at[headers[n], "amount1"] = material_list.at[
                                #                                               headers[n], "amount1"] + float(line[n])
                                if material_list.at[headers[n], "amount1"] < (liquid_limit-(liquid_limit*liquid_deadspace)):
                                    header_new = material_list.at[headers[n], "id1"]
                                else:
                                    #pass
                                    print("Refill liquid channels")
                                    exit()

                        if (material_list.at[headers[n], "type"]) == "solid":
                            #print(line[n])
                            if float(line[n]) > 0:
                                line[n] = float(line[n]) / 1000
                                #line[n] = float(line[n])
                            if pd.isnull(material_list.at[headers[n], "amount1"]):
                                if i == 29:
                                    material_list.at[headers[n], "amount1"] = (float(line[n]) * 3)
                                else:
                                    material_list.at[headers[n], "amount1"] = float(line[n])
                                #material_list.at[headers[n], "amount1"] = float(line[n])
                            else:
                                if i == 29:
                                    material_list.at[headers[n], "amount1"] = material_list.at[headers[n], "amount1"] + (float(line[n])*3)
                                else: material_list.at[headers[n], "amount1"] = material_list.at[headers[n], "amount1"] + float(line[n])
                                # material_list.at[headers[n], "amount1"] = material_list.at[
                                #                                               headers[n], "amount1"] + float(line[n])
                            header_new = material_list.at[headers[n], "id1"]
                        if (material_list.at[headers[n], "type"]) == "subsampling":

                            if SUBSAMPLE_SWITCH_THROUGH == False:

                                if pd.isnull(material_list.at[headers[n], "amount1"]):
                                    if i == 29:
                                        material_list.at[headers[n], "amount1"] = (float(line[n]) * 3)
                                    else:
                                        material_list.at[headers[n], "amount1"] = float(line[n])
                                    # material_list.at[headers[n], "amount1"] = float(line[n])
                                    header_new = material_list.at[headers[n], "id1"]
                                else:
                                    if i == 29:
                                        material_list.at[headers[n], "amount1"] = material_list.at[
                                                                                      headers[n], "amount1"] + (
                                                                                              float(line[n]) * 3)
                                    else:
                                        material_list.at[headers[n], "amount1"] = material_list.at[
                                                                                      headers[n], "amount1"] + float(
                                            line[n])
                                    # material_list.at[headers[n], "amount1"] = material_list.at[
                                    #                                               headers[n], "amount1"] + float(line[n])
                                    #if material_list.at[headers[n], "amount1"] < (
                                    #        liquid_limit - (liquid_limit * liquid_deadspace)):
                                    header_new = material_list.at[headers[n], "id1"]
                                    #else:
                                    #    pass
                                    #    print("Refill liquid channels")
                                       #exit()

                            if SUBSAMPLE_SWITCH_THROUGH == True:
                                for x in range(SUBSAMPLE_SIZE):
                                    #print(("amount{}").format(i+1))
                                    if (material_list.at[headers[n], ("amount{}").format(x+1)] + float(line[n])) > 36:
                                        if x == (SUBSAMPLE_SIZE - 1):
                                            header_new = "Invalid"
                                            print("Warning - Subsample plates need restocking")
                                            exit()
                                    continue

                                if WASH_ZEROS == True:
                                    wash_counter = wash_counter + wash_amount
                                    #print("Wash:" + material_list.at[headers[n], "Material"] + "\n For: " + wash_counter)

                                if WASH_ZEROS == False:
                                    if float(line[n]) > 0:
                                        wash_counter = wash_counter + wash_amount
                                        mat = headers[n]
                                        if not mat in in_wash:
                                            in_wash.append(mat)

                                if pd.isnull(material_list.at[headers[n], ("amount{}").format(x+1)]):
                                    material_list.at[headers[n], ("amount{}").format(x+1)] = float(line[n])
                                    header_new = material_list.at[headers[n], ("id{}").format(x + 1)]

                                    break

                                if (material_list.at[headers[n], ("amount{}").format(x+1)] + float(line[n])) <= 36:
                                    material_list.at[headers[n], ("amount{}").format(x + 1)] = material_list.at[headers[n], ("amount{}").format(x+1)] + float(line[n])
                                    header_new = material_list.at[headers[n], ("id{}").format(x + 1)]
                                    break













                    headers_dict["{}".format(line[0])].append(header_new)




                    headers_dict["{}".format(line[0])].append(float(line[n]))
                    # print(line[n])
                    # print((headers[n], n, material_list.at[headers[n], "type"]))
                    #print(line[n])
                    if material_list.at[headers[n], "type"] == "liquid":
                        headers_dict["{}".format(line[0])].append(sub_info_dict["liquid_dispenser"])
                    if material_list.at[headers[n], "type"] == "solid":
                        headers_dict["{}".format(line[0])].append(sub_info_dict["solid_dispenser"])
                    if material_list.at[headers[n], "type"] == "subsampling":
                        headers_dict["{}".format(line[0])].append(sub_info_dict["subsampling_dispenser"])

                if (material_list.at["Wash", "type"]) == "Wash solution":
                    if pd.isnull(material_list.at["Wash", ("amount{}").format(1)]):
                        material_list.at["Wash", ("amount{}").format(1)] = wash_counter
                    else:
                        material_list.at["Wash", ("amount{}").format(1)] = material_list.at[
                                                                               "Wash", ("amount{}").format(
                                                                                   1)] + wash_counter

                    if material_list.at["Wash", ("amount{}").format(1)] > (
                            subsample_wash_limit - (subsample_wash_limit * subsample_deadspace)):
                        print("Subsample wash bottle needs draining")
                        messagebox.showerror("Process error", "Subsample wash bottle needs draining")
                        exit()


                    #print(headers_dict)
                    #print(material_list[["amount1","amount2"]])

                    #print(material_list[headers[n]])
                    #print(headers[n])

                    #print(material_list)
                    #print(material_list.loc[headers[n], :])
                    #m = material_list.index.get_loc(headers[n])
                    #material_list.loc(m, :)
                    #print(material_list.index.get_loc(headers[n]))
                    #print(headers[n])



                    #"use df.at[row,col] for stuff it works"


                    #col += 1
                    #ws.cell(row=i + 3, column=col).value = float(line[n])
                    #col += 1


                    #working_df = working_df.append(headers_dict, ignore_index=True)
                    #working_df = working_df.append(headers_dict, ignore_index=True)
                    #print(working_df)

                    '''
                    assumes water is always the last addition - which it should be!
                    '''

                #     if headers[n] in liquids or headers[n] == headers[-1]:
                #         #ws.cell(row=i + 3, column=col).value = sub_info_dict["liquid_dispenser"]
                #         headers_dict["{}".format(line[0])].append(sub_info_dict["liquid_dispenser"])
                #     else:
                #         #ws.cell(row=i + 3, column=col).value = sub_info_dict["solid_dispenser"]
                #         headers_dict["{}".format(line[0])].append(sub_info_dict["solid_dispenser"])
                #     #col += 1
                # # set end params from col
                int_working_df = pd.DataFrame(data=headers_dict).transpose() # makes dataframes from the dicts, transposes
                #print("end: " + str(wash_counter))
                #ext_working_df = ext_working_df.append(int_working_df) # builds the experiment dataframe as the int gets reset
                #print(working_df)
                #print(int_working_df)
                  # puts batch_name+num into form code

                #print(int_working_df.head())


                col = 8  # sets col to H for value


                '''
                Here needs a check for running sample to add delay step, 21600s for 6hr
                Needs to end on col=9 
                
                Check run queue folder
                If 0, delay on 2nd plate
                If 1, delay on both plates
                max 4 plates in queue? (2 batches)
                
                '''

                # delay step
                # if i == 0:
                #     in_queue = self.check_run_queue()
                #     if in_queue == 0:
                #         ws.cell(row=i + 18, column=col).value = delay_time
                #     else:
                #         ws.cell(row=i + 3, column=col).value = delay_time
                #         ws.cell(row=i + 18, column=col).value = delay_time
                #col += 1

                ws.cell(row=i + 3, column=1).value = int_working_df.index[0]
                ws.cell(row=i + 3, column=2).value = int_working_df.index[0]
                ws.cell(row=i + 3, column=3).value = sub_info_dict["hazard_1"]
                ws.cell(row=i + 3, column=4).value = sub_info_dict["hazard_2"]
                ws.cell(row=i + 3, column=5).value = sub_info_dict["hazard_3"]
                n2 = 0

                '''
                Need to go back over and replace rows 14, 15 with data from 30? 
                n is column in df? 
                
                hard coded in to replace rows 14, 15 with the data from 30
                '''
                #REMOVE_ZEROS = True
                remove_zero_flag = False
                for n in range(len(int_working_df.columns)):

                    if REMOVE_ZEROS == True:
                        if remove_zero_flag == True:
                            remove_zero_flag = False
                            col += 1
                            continue

                        if int_working_df.at[int_working_df.index[0], n] == 0:
                            ws.cell(row=i + 3, column=col-1).value = None
                            remove_zero_flag = True
                            col += 1
                            continue




                    ws.cell(row=i + 3, column=col).value = int_working_df.at[int_working_df.index[0], n]

                        #print(int_working_df.at[int_working_df.index[0], n])
                    # if i == 29:
                    #     ws.cell(row=16, column=col).value = int_working_df.at[int_working_df.index[0], n]
                    #     ws.cell(row=17, column=col).value = int_working_df.at[int_working_df.index[0], n]
                    col += 1

                    # ws.cell(row=i + 3, column=col).value = int_working_df.at[int_working_df.index[0], n]
                    # n +=1
                    # col += 1
                    # ws.cell(row=i + 3, column=col).value = int_working_df.at[int_working_df.index[0], n]
                    # n +=1
                    # col += 1
                #col = col+len(int_working_df.columns)

                    #print(ws.cell(row=i + 3, column=col).value, int_working_df.at[int_working_df.index[0], n])
                # for n in range(len(line)):
                #     if headers[n] == "Name":
                #         continue
                #     ws.cell(row=i + 3, column=col).value = headers[n]
                #     #print(headers[n])
                #     col += 1
                #     ws.cell(row=i + 3, column=col).value = float(line[n])
                #     col += 1
                #
                #     '''
                #     assumes water is always the last addition - which it should be!
                #     '''
                #
                #     if headers[n] in liquids or headers[n] == headers[-1]:
                #         ws.cell(row=i + 3, column=col).value = sub_info_dict["liquid_dispenser"]
                #     else:
                #         ws.cell(row=i + 3, column=col).value = sub_info_dict["solid_dispenser"]
                #     col += 1
                if i == 29:
                    for x in range(col-7):
                        ws.cell(row=16, column=x + 8).value = ws.cell(row=31, column=x + 8).value
                        ws.cell(row=17, column=x + 8).value = ws.cell(row=31, column=x + 8).value



                #print(sub_info_dict["measurement_method"])

                ws.cell(row=i + 3, column=col).value = sub_info_dict["orbital_speed_rpm"]
                col += 1
                ws.cell(row=i + 3, column=col).value = sub_info_dict["orbital_id"]
                col += 1
                ws.cell(row=i + 3, column=col).value = sub_info_dict["illumination_time_secs"]
                col += 1
                ws.cell(row=i + 3, column=col).value = sub_info_dict["orbital_id"]
                col += 1
                try:
                    ws.cell(row=i + 3, column=col).value = sub_info_dict["measurement_method"]
                except:
                    pass



            #print(ext_working_df)

            #writer = pd.ExcelWriter(path, engine="openxyl")
            #ext_working_df.to_excel(excel_writer=ws, sheet_name="Formulations", engine="openpxyl", startcol=3, startrow=8, header=None, index=False)
            book.save(path)
            waste_file = open(self.directory_path + "Material tracking/waste.txt", "w")
            for w in in_wash:
                waste_file.write(w + "\n")

            waste_file.close()
            material_list.to_csv(self.directory_path + "Material tracking/material_list.csv")
            #print(ext_working_df)
        except IOError:
            print("Cannot create a batch file.")



    def submit_mini_batch(self, batch_name, quantities_list, liquids):
        '''
        Generates a new file for the workflow and puts it in the runqueue folder.
        E.g. quantities_list = [{'P10-HS' : 5, 'AscorbicAcid0-1M' : 4.99}, {'P10-HS' : 5.1, 'AscorbicAcid0-1M' : 4.39}]
        liquids = {'AscorbicAcid0-1M', 'NaCL-1-0M'}
        batch_name should be consistent with FE's direct upload experiment files for HOLLY
        The filename is set to batch_name + '.xlsx'


        '''
        try:

            path = self.directory_path + '/runqueue/' + batch_name + '.xlsx'
            book = load_workbook("fe_batch.xlsx")
            worksheet_names = book.sheetnames

            sub_info = self.submission_header.split("\n")
            sub_info_split = []
            for item in sub_info:
                sub_info_split.append(item.split(",")[0])
                sub_info_split.append(item.split(",")[1])

            sub_info_dict = dict(
                zip(sub_info_split[::2], sub_info_split[1::2]))  # makes dictionary of params in batch.temp

            book.active = worksheet_names.index("Experiment Details")  # sets params in exp. details sheet
            ws = book.active
            ws.cell(row=4, column=2).value = batch_name  # NEEDS TO BE SAME AS FILE NAME FOR TRACEABILITY
            ws.cell(row=5, column=2).value = sub_info_dict["objective"]
            ws.cell(row=11, column=2).value = sub_info_dict["o2_level"]
            ws.cell(row=12, column=2).value = sub_info_dict["cap_vials"]

            book.active = worksheet_names.index(
                "Formulations")  # sets working sheet to formulations to print out amounts
            ws = book.active

            for i, quantity in enumerate(quantities_list):
                #print(i, quantity)
                line = self.submission_line

                #############################################
                # JUST ADD WATER. REMOVE FOR OTHER EXPERIMENTS
                water = 5

                # TEST
                # test = 10

                for chem, amount in quantity.items():
                    if chem in liquids:
                        water = water - amount

                    # TEST
                    # test = test - (0.5 - amount)**2

                if water < 0:
                    print("LAST CHINESE WARNING! The constraints did not work. Total volume is more than 5ml.")
                    water = 0
                #############################################

                for chem, amount in quantity.items():
                    line = self.patterns[chem].sub(self.FLOAT_FORMAT.format(amount), line)

                line = self.patterns['idx'].sub(str(i), line)
                line = self.patterns["batch_name"].sub(batch_name, line)
                line = self.patterns['sample_number'].sub(str(i + 1), line)

                #############################################
                # JUST ADD WATER. REMOVE FOR OTHER EXPERIMENTS
                line = self.water.sub(self.FLOAT_FORMAT.format(water), line)

                # TEST
                # line = self.test.sub(self.FLOAT_FORMAT.format(test), line)
                #############################################

                # f.write(line)
                line = line.split(",")
                headers = self.submission_compounds.split(",")
                line = line[2:]
                headers = headers[2:]  # trims id and number as not required.
                row = i + 3

                ws.cell(row=i + 3, column=1).value = line[0]  # puts batch_name+num into form code
                ws.cell(row=i + 3, column=2).value = line[
                    0]  # puts batch name+num into form des until HOLLY supports form codes
                # ws.cell(row=i + 3, column=2).value = sub_info_dict["description"]

                ws.cell(row=i + 3, column=3).value = sub_info_dict["hazard_1"]
                ws.cell(row=i + 3, column=4).value = sub_info_dict["hazard_2"]
                ws.cell(row=i + 3, column=5).value = sub_info_dict["hazard_3"]
                col = 8  # sets col to H
                for n in range(len(line)):
                    if headers[n] == "Name":
                        continue
                    ws.cell(row=i + 3, column=col).value = headers[n]
                    #print(headers[n])
                    col += 1
                    ws.cell(row=i + 3, column=col).value = float(line[n])
                    col += 1

                    '''
                    assumes water is always the last addition - which it should be!
                    '''

                    if headers[n] in liquids or headers[n] == headers[-1]:
                        ws.cell(row=i + 3, column=col).value = sub_info_dict["liquid_dispenser"]
                    else:
                        ws.cell(row=i + 3, column=col).value = sub_info_dict["solid_dispenser"]
                    col += 1
                # set end params from col




                ws.cell(row=i + 3, column=col).value = sub_info_dict["orbital_speed_rpm"]
                col += 1
                ws.cell(row=i + 3, column=col).value = sub_info_dict["orbital_id"]
                col += 1
                ws.cell(row=i + 3, column=col).value = sub_info_dict["illumination_time_secs"]
                col += 1
                ws.cell(row=i + 3, column=col).value = sub_info_dict["orbital_id"]
                col += 1
                ws.cell(row=i + 3, column=col).value = sub_info_dict["measurement_method"]

                # f.write('\n')
            book.save(path)
        # try:
        #     path = self.directory_path+'/runqueue/' + batch_name + '.run'
        #     with open(path,"w") as f:
        #         f.write(f'batch_name:{batch_name}\n')
        #
        #
        #
        #
        #         f.write(self.submission_header + '\n\n')
        #
        #         f.write(datetime.datetime.now().strftime('submit_start_datetime:%Y.%m.%d.%H.%M.%S\n\n'))
        #
        #         f.write(self.submission_compounds) #table headers
        #         print(self.submission_compounds)
        #
        #         for i, quantity in enumerate(quantities_list):
        #             # print(i, quantity)
        #             line = self.submission_line
        #              #terms yet to be assigned values
        #             #############################################
        #             #JUST ADD WATER. REMOVE FOR OTHER EXPERIMENTS
        #             water = 5
        #
        #             #TEST
        #             # test = 10
        #
        #             for chem, amount in quantity.items():
        #                 if chem in liquids :
        #                     water = water - amount
        #
        #                 #TEST
        #                 # test = test - (0.5 - amount)**2
        #
        #             if water < 0:
        #                 print("LAST CHINESE WARNING! The constraints did not work. Total volume is more than 5ml.")
        #                 water = 0
        #             #############################################
        #
        #
        #             for chem, amount in quantity.items():
        #                 line = self.patterns[chem].sub(self.FLOAT_FORMAT.format(amount), line) #fills in chemical amounts
        #             line = self.patterns['idx'].sub(str(i), line)
        #              #idx assigned
        #             line = self.patterns["batch_name"].sub(batch_name, line)
        #
        #             line = self.patterns['sample_number'].sub(str(i+1), line)
        #
        #
        #             #############################################
        #             #JUST ADD WATER. REMOVE FOR OTHER EXPERIMENTS
        #             line = self.water.sub(self.FLOAT_FORMAT.format(water), line)
        #
        #             #TEST
        #             # line = self.test.sub(self.FLOAT_FORMAT.format(test), line)
        #             #############################################
        #
        #             f.write(line)
        #
        #             f.write('\n')

        except IOError:
            print("Cannot create a batch file.")

if __name__ == "__main__":
    from experiment import Experiment
    exp = Experiment()
    print(exp.rng.keys())
    parser = Parser(list(exp.rng.keys()),'./example_exploratory/')
    parser.process_completed_file("gC3N4-16.run")
    test_submit = [{'P10-HS' : 5, 'AscorbicAcid0-1M' : 4.99}, {'P10-HS' : 5.1, 'AscorbicAcid0-1M' : 4.39}]
    parser.submit_mini_batch(batch_name='test', quantities_list = test_submit)